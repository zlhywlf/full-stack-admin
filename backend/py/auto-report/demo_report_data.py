import sqlite3
import pathlib
from sqlite3 import Connection
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def get_path(name: str):
    """
    获取数据文件绝对路径
    :param name: 文件名称
    :return: 文件绝对路径
    """
    return str(pathlib.Path.home()) + r'/Desktop/data/' + name


def read_sql(name: str) -> str:
    """
    读取 sql 脚本文件
    :param name: 脚本名称
    :return: sql 脚本字符串
    """
    with open('./data/%s' % name, 'r') as f:
        return f.read()


def get_insert_sql(table_name: str, keys: tuple) -> str:
    """
    组装插入数据 sql
    :param table_name: 插入表名
    :param keys: 列名
    :return: 插入数据 sql 字符串
    """
    columns = ','.join(keys)
    placeholders = ','.join('?' * len(keys))
    return 'INSERT INTO {table_name} ({columns}) VALUES ({placeholders})'.format(table_name=table_name,
                                                                                 columns=columns,
                                                                                 placeholders=placeholders)


def insert_data(connection: Connection, source_name: str, table_name: str) -> None:
    """
    将 excel 数据写入数据库，首行为标题的一维数据
    :param connection: 数据库连接
    :param source_name: excel 数据源
    :param table_name: 写入数据库表名
    :return: None
    """
    wb = load_workbook(get_path(source_name), True)
    ws: Worksheet = wb.active
    titles = tuple()
    print('*' * 10, table_name, 'start', '*' * 10)
    for index, row in enumerate(ws.values):
        if index == 0:
            titles = row
            print('读取标题:', row)
        else:
            with connection:
                connection.execute(get_insert_sql(table_name, titles), row)
    print('写入记录:', (ws.max_row - 1))
    print('*' * 10, table_name, 'end  ', '*' * 10)


if __name__ == '__main__':
    # 1 连接数据库
    conn = sqlite3.connect(r'./data/data.db')

    # 2 创建表
    conn.executescript(read_sql(r'create.sql'))

    # 3 插入数据
    insert_data(conn, r'dataset.xlsx', r'student')

    # 4 关闭连接
    conn.close()
