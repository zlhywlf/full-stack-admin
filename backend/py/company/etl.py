import os
import re
import sqlite3
from openpyxl import load_workbook
from sqlite3 import Connection, Cursor
from pyecharts import options as opts
from pyecharts.charts import Line, Page
from pyecharts.components import Table
from pyecharts.globals import CurrentConfig


def get_file_path(name: str):
    return os.getcwd() + r'/dist/' + name


def get_data_path(name: str):
    return os.getcwd() + r'/backend/py/company/data/' + name


def get_chinese(text: str) -> str:
    return ''.join(re.findall(r'[\u4e00-\u9fa5]+', text)).strip()


def get_value(v: any):
    if isinstance(v, str):
        if v.strip() != '':
            print(v)
    return 0 if isinstance(v, str) else v if v else 0


def get_data(table_name: str, connection: Connection):
    wb = load_workbook(get_file_path(table_name), True)
    for name in wb.sheetnames:
        if '.' in name:
            print(name)
            year, month = name.split('.')
            ws = wb[name]
            row_num = 16
            while True:
                # 工序
                p = get_chinese(ws.cell(row_num, 2).value)
                if ('系数' in p):
                    break
                data = [year, month]
                data.append(p)
                # 产品
                data.append(get_value(ws.cell(row_num+1, 2+1).value))
                # 燃料消耗
                for col_num in range(5, 15):
                    data.append(get_value(ws.cell(row_num+1, col_num).value))
                # 能源消耗
                for col_num in range(19, 26):
                    data.append(get_value(ws.cell(row_num+1, col_num).value))
                row_num += 3
                with connection:
                    connection.execute(get_insert_sql("data", ('年', '月', '工序', '产品', '无烟煤', '贫瘦煤',
                                                               '烟煤', '冶金焦', '焦丁', '柴油', '汽油', '焦面', '高炉煤气', '转炉煤气', '电', '蒸汽', '新水', '氧气', '氮气', '氩气', '压缩空气')), data)
    wb.close


def read_sql(name: str) -> str:
    with open(get_data_path(name), 'r') as f:
        return f.read()


def get_insert_sql(table_name: str, keys: tuple) -> str:
    columns = ','.join(keys)
    placeholders = ','.join('?' * len(keys))
    return 'INSERT INTO {table_name} ({columns}) VALUES ({placeholders})'.format(table_name=table_name,
                                                                                 columns=columns,
                                                                                 placeholders=placeholders)


def do_etl(conn: Connection):
    conn.executescript(read_sql(r'create.sql'))
    for n in [r"磐石建龙2021年能源报表.xlsx", r"磐石建龙2022年能源报表.xlsx"]:
        get_data(n, conn)


def dict_factory(cursor, row):
    d = {}
    for idx, col in enumerate(cursor.description):
        d[col[0]] = row[idx]
    return d


def get_toolboxOpts():
    return opts.ToolboxOpts(is_show=True, feature=opts.ToolBoxFeatureOpts(
        data_zoom=opts.ToolBoxFeatureDataZoomOpts(False),
        magic_type=opts.ToolBoxFeatureMagicTypeOpts(False),
        restore=opts.ToolBoxFeatureRestoreOpts(False)
    ))


def handle_line(title: str, unit: str, res: list) -> Line:
    line = Line(opts.InitOpts(page_title=title,
                width="100%", height="300px"))
    is_first = True
    for i in res:
        if is_first:
            line.add_xaxis(list(map(lambda x: x, i['月'].split(','))))
            is_first = False
        line.add_yaxis(i['年'], list(
            map(lambda x: float(x), i['产品'].split(','))), is_smooth=True, label_opts=opts.LabelOpts(is_show=False))

    line.set_global_opts(
        tooltip_opts=opts.TooltipOpts(trigger="axis"),
        toolbox_opts=get_toolboxOpts(),
        xaxis_opts=opts.AxisOpts(
            type_="category", splitline_opts=opts.SplitLineOpts(False)),
        yaxis_opts=opts.AxisOpts(splitline_opts=opts.SplitLineOpts(False)),
        title_opts=opts.TitleOpts(subtitle='单位: '+unit),
    )
    return line


def handle_production(p: Page, c: Cursor):
    tb = [
        {
            'tb': '烧结机',
            'name': '烧结矿产量',
            'unit': '万吨'
        },
        {
            'tb': '竖炉',
            'name': '球团矿产量',
            'unit': '万吨'
        },
        {
            'tb': '炼铁合计',
            'name': '合格铁产量',
            'unit': '万吨'
        },
        {
            'tb': '炼钢',
            'name': '钢坯产量',
            'unit': '万吨'
        },
        {
            'tb': '轧钢',
            'name': '窄带钢产量',
            'unit': '万吨'
        },
        {
            'tb': '白灰窑',
            'name': '石灰产量',
            'unit': '万吨'
        },
        {
            'tb': '氧气',
            'name': '氧气产量',
            'unit': '万立方米'
        },
        {
            'tb': '氮气',
            'name': '氮气产量',
            'unit': '万立方米'
        },
        {
            'tb': '新水',
            'name': '新水产量',
            'unit': '万立方米'
        },
        {
            'tb': '发电',
            'name': '发电产量',
            'unit': '万千瓦时'
        }
    ]
    base_sql = read_sql(r"产量.sql")
    for t in tb:
        res = c.execute(base_sql.format(t['tb'])).fetchall()
        p.add(handle_line(t['name'], t['unit'], res))


def handle_fuel(p_name: str, p: Page, c: Cursor, tb: list):
    base_sql = read_sql(r"燃料.sql")
    for t in tb:
        res = c.execute(base_sql.format(t, p_name)).fetchall()
        p.add(handle_line(p_name+'-'+t+'消耗', '万吨标煤', res))


if __name__ == '__main__':
    conn = sqlite3.connect(get_data_path(r'data.db'))

    # do_etl(conn)

    cursor = conn.cursor()
    cursor.row_factory = dict_factory

    workdir = os.getcwd()
    CurrentConfig.GLOBAL_ENV.loader.searchpath.append(
        workdir+"/backend/py/dashboard")
    page = Page(page_title="2021-2022年磐石建龙能源报表", is_remove_br=True)

    page.add(Table(page_title='a报表1').add(['报表1'], [['报表1']]),
             Table(page_title='b报表1').add(['报表1'], [['报表1']]),
             Table(page_title='a报表2').add(['报表2'], [['报表2']]),
             Table(page_title='b报表2').add(['报表2'], [['报表2']]),
             Table(page_title='a报表3').add(['报表3'], [['报表3']]),
             Table(page_title='b报表3').add(['报表3'], [['报表3']]),)
    handle_production(page, cursor)
    # tb = ['无烟煤', '贫瘦煤', '烟煤', '冶金焦', '焦丁', '柴油', '汽油', '焦面','高炉煤气', '转炉煤气', '电', '蒸汽', '新水', '氧气', '氮气', '氩气', '压缩空气']
    handle_fuel('烧结机', page, cursor, [
                '无烟煤',  '焦面', '高炉煤气', '转炉煤气', '电', '蒸汽'])
    handle_fuel('竖炉', page, cursor, ['高炉煤气',  '电'])
    handle_fuel('炼铁合计', page, cursor, [
                '无烟煤', '贫瘦煤', '烟煤', '冶金焦', '焦丁',  '高炉煤气',  '电', '蒸汽', '新水', '氧气', '氮气', ])
    path = workdir+"/dist/mycharts.html"
    page.render(path, "template.html")
    conn.close()
